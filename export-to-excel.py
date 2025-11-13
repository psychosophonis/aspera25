#!/usr/bin/env python3
"""
Export with Participants sheet as single source of truth
Other sheets reference by name only
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def style_worksheet(ws, header_color="4472C4"):
    """Apply clean styling to worksheet"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(60, max(12, max_length + 2))
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

def main():
    src = Path("conference-data.json")
    out = Path("conference-data.xlsx")
    
    with open(src, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # =================================================================
    # PARTICIPANTS: Extract all unique people with their institutions
    # =================================================================
    participants = {}  # name -> institution (most complete)
    
    def add_participant(person_obj):
        """Add person to registry, keeping most complete institution"""
        name = person_obj['name']
        inst = person_obj.get('institution')
        # Keep most complete institution (non-null takes priority)
        if name not in participants or (inst and not participants[name]):
            participants[name] = inst
    
    # From sessions
    for session in data.get('sessions', {}).values():
        for p in session.get('presenters', []):
            add_participant(p)
        for paper in session.get('papers', []):
            for a in paper.get('authors', []):
                add_participant(a)
    
    # From events
    for event in data.get('events', []):
        if 'presenters' in event and isinstance(event['presenters'], list):
            for p in event['presenters']:
                add_participant(p)
        
        # From screening films
        if event.get('type') == 'screening' and 'films' in event:
            for film in event.get('films', []):
                for creative in film.get('creatives', []):
                    add_participant(creative)
    
    participants_df = pd.DataFrame([
        {'Name': name, 'Institution': inst}
        for name, inst in sorted(participants.items())
    ])
    
    # =================================================================
    # SESSIONS & PAPERS: Names only, no institutions
    # =================================================================
    session_papers = []
    
    for sid, session in sorted(data.get('sessions', {}).items(), key=lambda x: int(x[0])):
        presenters = '; '.join([p['name'] for p in session.get('presenters', [])])
        
        session_papers.append({
            'Session #': f"SESSION {sid}",
            'Type': session.get('type', ''),
            'Title': session.get('title', ''),
            'Description': session.get('description', ''),
            'Presenters/Authors': presenters,
            'Abstract': ''
        })
        
        for paper in session.get('papers', []):
            authors = '; '.join([a['name'] for a in paper.get('authors', [])])
            
            session_papers.append({
                'Session #': '',
                'Type': 'paper',
                'Title': paper.get('title', ''),
                'Description': '',
                'Presenters/Authors': authors,
                'Abstract': paper.get('abstract', '')
            })
    
    sessions_df = pd.DataFrame(session_papers)
    
    # =================================================================
    # EVENTS: Names only
    # =================================================================
    event_rows = []
    
    for event in data.get('events', []):
        presenters = ''
        if 'presenters' in event and isinstance(event['presenters'], list):
            presenters = '; '.join([p['name'] for p in event['presenters']])
        
        event_rows.append({
            'Day': event.get('day', ''),
            'Time': event.get('time', ''),
            'Venue': event.get('venue', ''),
            'Type': event.get('type', ''),
            'Content': event.get('content', ''),
            'Presenters': presenters,
            'Abstract': event.get('abstract', '') if event.get('type') == 'plenary' else ''
        })
    
    events_df = pd.DataFrame(event_rows)
    
    # =================================================================
    # SCREENINGS: Names only for creatives
    # =================================================================
    screening_rows = []
    
    for event in data.get('events', []):
        if event.get('type') == 'screening':
            screening_rows.append({
                'Screening Slot': f"{event.get('day', '')} - {event.get('time', '')} - {event.get('venue', '')}",
                'Film Title': '',
                'Year': '',
                'Duration (mins)': '',
                'Creatives': ''
            })
            
            for film in event.get('films', []):
                creatives = '; '.join([c['name'] for c in film.get('creatives', [])])
                
                screening_rows.append({
                    'Screening Slot': '',
                    'Film Title': film.get('title', ''),
                    'Year': film.get('year', ''),
                    'Duration (mins)': film.get('duration', ''),
                    'Creatives': creatives
                })
    
    screenings_df = pd.DataFrame(screening_rows)
    
    # =================================================================
    # Write to Excel
    # =================================================================
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        participants_df.to_excel(writer, sheet_name='Participants', index=False)
        sessions_df.to_excel(writer, sheet_name='Sessions & Papers', index=False)
        events_df.to_excel(writer, sheet_name='Events', index=False)
        screenings_df.to_excel(writer, sheet_name='Screenings', index=False)
    
    wb = load_workbook(out)
    style_worksheet(wb['Participants'], header_color="4472C4")
    style_worksheet(wb['Sessions & Papers'], header_color="70AD47")
    style_worksheet(wb['Events'], header_color="FFC000")
    style_worksheet(wb['Screenings'], header_color="C55A11")
    wb.save(out)
    
    print(f"✅ Excel workbook created: {out.resolve()}")
    print(f"\n📊 Summary:")
    print(f"   • {len(participants)} participants")
    print(f"   • {len(data.get('sessions', {}))} sessions with {sum(len(s.get('papers', [])) for s in data.get('sessions', {}).values())} papers")
    print(f"   • {len(data.get('events', []))} events")
    print(f"   • {len([e for e in data.get('events', []) if e.get('type') == 'screening'])} screening slots")
    print(f"\n💡 Workflow:")
    print(f"   1. Edit institutions in Participants sheet (single source of truth)")
    print(f"   2. Edit names/content in other sheets (just names, no institutions)")
    print(f"   3. Import: institutions are looked up from Participants sheet")

if __name__ == "__main__":
    main()